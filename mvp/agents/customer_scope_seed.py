"""Helpers for generating mock Databricks seed data from customer scope samples."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_DEFAULT_SCOPE_WORKBOOK = Path(__file__).resolve().parent.parent / "customer_input" / "accont_scope_query_result.csv"
_WORKBOOK_SHEET_NAME = "result"
_MOCK_EMAIL_ALIASES = {
    "DaichiM@M365CPI89838450.OnMicrosoft.com": "bartek.niezabitowski@veeam.com",
}


@dataclass(frozen=True)
class MockCustomerSeedDataset:
    accounts: list[dict[str, Any]]
    object_territory_associations: list[dict[str, Any]]
    territories: list[dict[str, Any]]
    user_territory_associations: list[dict[str, Any]]
    users: list[dict[str, Any]]
    aiq_rows: list[dict[str, Any]]
    contacts: list[dict[str, Any]]


def default_scope_workbook_path() -> Path:
    return _DEFAULT_SCOPE_WORKBOOK


def resolve_scope_workbook_path(raw_path: str | Path | None = None) -> Path:
    configured = Path(raw_path).expanduser() if raw_path else default_scope_workbook_path()
    if configured.is_absolute():
        return configured
    return (Path.cwd() / configured).resolve()


def _normalize_text(value: Any) -> str | None:
    normalized = str(value or "").strip()
    if not normalized or normalized.lower() == "null":
        return None
    return normalized


def _stable_digest(*parts: str) -> str:
    seed = "|".join(part.strip().lower() for part in parts)
    return hashlib.sha256(seed.encode("utf-8")).hexdigest()


def _stable_int(seed: str, minimum: int, maximum: int) -> int:
    span = maximum - minimum + 1
    return minimum + (int(seed[:8], 16) % span)


def _stable_float(seed: str, minimum: float, maximum: float) -> float:
    basis = int(seed[:8], 16) / 0xFFFFFFFF
    return round(minimum + ((maximum - minimum) * basis), 2)


def _stable_bool(seed: str, offset: int = 0) -> bool:
    return bool((int(seed[offset: offset + 2], 16) % 2))


def _slugify_account_name(name: str) -> str:
    lowered = "".join(ch.lower() if ch.isalnum() else "-" for ch in name)
    compact = "-".join(part for part in lowered.split("-") if part)
    return compact[:32] or "account"


def _territory_id(name: str, index: int) -> str:
    digest = _stable_digest(name)
    return f"0TT{index:03d}{digest[:9].upper()}"


def _fallback_user_id(email: str, index: int) -> str:
    digest = _stable_digest(email)
    return f"005MOCK{index:03d}{digest[:8].upper()}"


def load_scope_workbook_rows(path: str | Path | None = None) -> list[dict[str, str | None]]:
    workbook_path = resolve_scope_workbook_path(path)
    raw_bytes = workbook_path.read_bytes()
    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    sheet_name = _WORKBOOK_SHEET_NAME if _WORKBOOK_SHEET_NAME in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value).strip() if value is not None else "" for value in next(rows)]
    dataset: list[dict[str, str | None]] = []
    for row in rows:
        record = {
            header[index]: _normalize_text(value)
            for index, value in enumerate(row[: len(header)])
            if header[index]
        }
        if record:
            dataset.append(record)
    return dataset


def build_mock_customer_seed_dataset(rows: list[dict[str, str | None]]) -> MockCustomerSeedDataset:
    accounts_by_id: dict[str, dict[str, Any]] = {}
    territories_by_name: dict[str, dict[str, Any]] = {}
    users_by_id: dict[str, dict[str, Any]] = {}
    object_territory_associations: set[tuple[str, str]] = set()
    user_territory_associations: set[tuple[str, str]] = set()
    aiq_rows_by_key: dict[tuple[str, str], dict[str, Any]] = {}

    for row_index, row in enumerate(rows, start=1):
        account_id = _normalize_text(row.get("vpower_account_id"))
        account_name = _normalize_text(row.get("vpower_account_name"))
        sales_team = _normalize_text(row.get("salesteam"))
        if not account_id or not account_name or not sales_team:
            continue

        legacy_id = _normalize_text(row.get("legacy_account_id"))
        parent_id = _normalize_text(row.get("vpower_parent_account_id"))
        parent_name = _normalize_text(row.get("vpower_parent_account_name"))
        user_id = _normalize_text(row.get("UserId"))
        email = _normalize_text(row.get("Email"))

        territory = territories_by_name.get(sales_team)
        if territory is None:
            territory = {
                "Id": _territory_id(sales_team, len(territories_by_name) + 1),
                "Name": sales_team,
                "Territory2ModelId": "0MAcx000000Arz7GAC",
                "Territory2TypeId": "0M5cx0000000E2zCAE",
                "vdm_is_hard_deleted": False,
                "__END_AT": None,
            }
            territories_by_name[sales_team] = territory

        if email:
            normalized_user_id = user_id or _fallback_user_id(email, len(users_by_id) + 1)
            users_by_id.setdefault(
                normalized_user_id,
                {
                    "Id": normalized_user_id,
                    "Email": email,
                    "vdm_is_hard_deleted": False,
                    "__END_AT": None,
                },
            )
            user_territory_associations.add((normalized_user_id, territory["Id"]))

        accounts_by_id.setdefault(
            account_id,
            {
                "Id": account_id,
                "RCA_AccountMigrationExternalId__c": legacy_id,
                "Name": account_name,
                "ParentId": parent_id,
                "IsDeleted": False,
                "vdm_is_hard_deleted": False,
                "__END_AT": None,
            },
        )
        object_territory_associations.add((account_id, territory["Id"]))

        if parent_id and parent_name:
            accounts_by_id.setdefault(
                parent_id,
                {
                    "Id": parent_id,
                    "RCA_AccountMigrationExternalId__c": None,
                    "Name": parent_name,
                    "ParentId": None,
                    "IsDeleted": False,
                    "vdm_is_hard_deleted": False,
                    "__END_AT": None,
                },
            )

        planner_account_id = legacy_id or account_id
        aiq_key = (planner_account_id, sales_team)
        if aiq_key not in aiq_rows_by_key:
            digest = _stable_digest(planner_account_id, sales_team)
            company_name = parent_name or account_name
            aiq_rows_by_key[aiq_key] = {
                "account_id": planner_account_id,
                "account_name": account_name,
                "company_name": company_name,
                "sales_team": sales_team,
                "xf_score_previous_day": _stable_float(digest, 61.0, 96.0),
                "xf_score_diff_pct": _stable_float(digest[8:], 1.0, 18.0),
                "intent": _stable_float(digest[16:], 67.0, 98.0),
                "competitive": _stable_float(digest[24:], 0.0, 12.0),
                "upsell": _stable_float(digest[32:], 0.0, 88.0),
                "fit": _stable_int(digest[4:], 55, 95),
                "need": _stable_int(digest[12:], 48, 92),
                "vdp_why": f"{account_name} shows active infrastructure and resilience buying signals.",
                "kasten_why": f"{company_name} has container and modernization indicators worth qualifying.",
                "o365_why": f"{company_name} appears to have Microsoft 365 protection whitespace.",
                "vbsf_why": f"{account_name} shows SaaS workflow backup relevance.",
                "cloud_why": f"{company_name} has cloud workload growth with backup expansion potential.",
                "sales_play_sell_vdp": _stable_bool(digest, 0),
                "sales_play_sell_kasten": _stable_bool(digest, 2),
                "sales_play_sell_o365": _stable_bool(digest, 4),
                "sales_play_sell_vbsf": _stable_bool(digest, 6),
                "sales_play_sell_cloud": _stable_bool(digest, 8),
                "sales_play_sell_vault": _stable_bool(digest, 10),
                "sales_play_vmware_migration": _stable_bool(digest, 12),
                "sales_play_upsell_vdp": _stable_bool(digest, 14),
                "sales_play_convert_to_vdc": _stable_bool(digest, 16),
            }

    contacts: list[dict[str, Any]] = []
    unique_accounts_for_contacts = sorted(
        aiq_rows_by_key.values(),
        key=lambda row: (str(row["sales_team"]).lower(), str(row["account_name"]).lower()),
    )
    for index, aiq_row in enumerate(unique_accounts_for_contacts, start=1):
        slug = _slugify_account_name(str(aiq_row["account_name"]))
        contacts.append(
            {
                "domain_account_id": aiq_row["account_id"],
                "first_name": "Primary",
                "last_name": f"Contact{index}",
                "name": f"Primary Contact {index}",
                "title": "Director of Infrastructure",
                "job_position": "IT Leaders",
                "email": f"primary-contact-{index}@{slug}.example",
                "phone": f"555-{1000 + index:04d}",
                "engagement_level": "Engaged",
                "contact_stage": "Marketing Qualified Contact",
                "last_activity_date": None,
                "do_not_call": False,
            }
        )

    user_email_to_id = {str(row["Email"]).lower(): str(row["Id"]) for row in users_by_id.values()}
    territory_ids_by_user_id: dict[str, set[str]] = {}
    for user_id, territory_id in user_territory_associations:
        territory_ids_by_user_id.setdefault(user_id, set()).add(territory_id)

    for alias_email, source_email in _MOCK_EMAIL_ALIASES.items():
        source_user_id = user_email_to_id.get(source_email.lower())
        if not source_user_id:
            continue
        alias_user_id = _fallback_user_id(alias_email, len(users_by_id) + 1)
        users_by_id.setdefault(
            alias_user_id,
            {
                "Id": alias_user_id,
                "Email": alias_email,
                "vdm_is_hard_deleted": False,
                "__END_AT": None,
            },
        )
        for territory_id in sorted(territory_ids_by_user_id.get(source_user_id, set())):
            user_territory_associations.add((alias_user_id, territory_id))

    return MockCustomerSeedDataset(
        accounts=sorted(accounts_by_id.values(), key=lambda row: str(row["Id"]).lower()),
        object_territory_associations=[
            {
                "ObjectId": object_id,
                "Territory2Id": territory_id,
                "IsDeleted": False,
                "vdm_is_hard_deleted": False,
                "__END_AT": None,
            }
            for object_id, territory_id in sorted(object_territory_associations)
        ],
        territories=sorted(territories_by_name.values(), key=lambda row: str(row["Name"]).lower()),
        user_territory_associations=[
            {
                "Territory2Id": territory_id,
                "UserId": user_id,
                "vdm_is_hard_deleted": False,
                "__END_AT": None,
            }
            for user_id, territory_id in sorted(user_territory_associations)
        ],
        users=sorted(users_by_id.values(), key=lambda row: str(row["Email"]).lower()),
        aiq_rows=sorted(
            aiq_rows_by_key.values(),
            key=lambda row: (str(row["sales_team"]).lower(), str(row["account_name"]).lower()),
        ),
        contacts=contacts,
    )


def _sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    escaped = str(value).replace("'", "''")
    return f"'{escaped}'"


def _render_insert(table_fqn: str, columns: list[str], rows: list[dict[str, Any]], *, overwrite: bool) -> str:
    if not rows:
        return ""
    verb = "INSERT OVERWRITE" if overwrite else "INSERT INTO"
    values_sql = []
    for row in rows:
        values_sql.append(
            "(" + ", ".join(_sql_literal(row.get(column)) for column in columns) + ")"
        )
    return f"{verb} {table_fqn} VALUES\n  " + ",\n  ".join(values_sql) + ";"


def render_mock_customer_seed_sql(
    rows: list[dict[str, str | None]],
    *,
    catalog_placeholder: str = "__CATALOG__",
) -> str:
    dataset = build_mock_customer_seed_dataset(rows)
    account_table = f"{catalog_placeholder}.sf_vpower_bronze.account"
    object_territory_table = f"{catalog_placeholder}.sf_vpower_bronze.objectterritory2association"
    territory_table = f"{catalog_placeholder}.sf_vpower_bronze.territory2"
    user_territory_table = f"{catalog_placeholder}.sf_vpower_bronze.userterritory2association"
    user_table = f"{catalog_placeholder}.sf_vpower_bronze.`user`"
    aiq_table = f"{catalog_placeholder}.data_science_account_iq_gold.account_iq_scores"
    contact_table = f"{catalog_placeholder}.account_iq_gold.aiq_contact"

    statements = [
        f"CREATE SCHEMA IF NOT EXISTS {catalog_placeholder}.sf_vpower_bronze;",
        f"""
CREATE OR REPLACE TABLE {account_table} (
  Id STRING NOT NULL,
  RCA_AccountMigrationExternalId__c STRING,
  Name STRING NOT NULL,
  ParentId STRING,
  IsDeleted BOOLEAN,
  vdm_is_hard_deleted BOOLEAN,
  `__END_AT` TIMESTAMP
);
""".strip(),
        f"""
CREATE OR REPLACE TABLE {object_territory_table} (
  ObjectId STRING NOT NULL,
  Territory2Id STRING NOT NULL,
  IsDeleted BOOLEAN,
  vdm_is_hard_deleted BOOLEAN,
  `__END_AT` TIMESTAMP
);
""".strip(),
        f"""
CREATE OR REPLACE TABLE {territory_table} (
  Id STRING NOT NULL,
  Name STRING NOT NULL,
  Territory2ModelId STRING NOT NULL,
  Territory2TypeId STRING NOT NULL,
  vdm_is_hard_deleted BOOLEAN,
  `__END_AT` TIMESTAMP
);
""".strip(),
        f"""
CREATE OR REPLACE TABLE {user_territory_table} (
  Territory2Id STRING NOT NULL,
  UserId STRING NOT NULL,
  vdm_is_hard_deleted BOOLEAN,
  `__END_AT` TIMESTAMP
);
""".strip(),
        f"""
CREATE OR REPLACE TABLE {user_table} (
  Id STRING NOT NULL,
  Email STRING NOT NULL,
  vdm_is_hard_deleted BOOLEAN,
  `__END_AT` TIMESTAMP
);
""".strip(),
        _render_insert(
            account_table,
            ["Id", "RCA_AccountMigrationExternalId__c", "Name", "ParentId", "IsDeleted", "vdm_is_hard_deleted", "__END_AT"],
            dataset.accounts,
            overwrite=True,
        ),
        _render_insert(
            object_territory_table,
            ["ObjectId", "Territory2Id", "IsDeleted", "vdm_is_hard_deleted", "__END_AT"],
            dataset.object_territory_associations,
            overwrite=True,
        ),
        _render_insert(
            territory_table,
            ["Id", "Name", "Territory2ModelId", "Territory2TypeId", "vdm_is_hard_deleted", "__END_AT"],
            dataset.territories,
            overwrite=True,
        ),
        _render_insert(
            user_territory_table,
            ["Territory2Id", "UserId", "vdm_is_hard_deleted", "__END_AT"],
            dataset.user_territory_associations,
            overwrite=True,
        ),
        _render_insert(
            user_table,
            ["Id", "Email", "vdm_is_hard_deleted", "__END_AT"],
            dataset.users,
            overwrite=True,
        ),
        _render_insert(
            aiq_table,
            [
                "account_id",
                "account_name",
                "company_name",
                "sales_team",
                "xf_score_previous_day",
                "xf_score_diff_pct",
                "intent",
                "competitive",
                "upsell",
                "fit",
                "need",
                "vdp_why",
                "kasten_why",
                "o365_why",
                "vbsf_why",
                "cloud_why",
                "sales_play_sell_vdp",
                "sales_play_sell_kasten",
                "sales_play_sell_o365",
                "sales_play_sell_vbsf",
                "sales_play_sell_cloud",
                "sales_play_sell_vault",
                "sales_play_vmware_migration",
                "sales_play_upsell_vdp",
                "sales_play_convert_to_vdc",
            ],
            dataset.aiq_rows,
            overwrite=False,
        ),
        _render_insert(
            contact_table,
            [
                "domain_account_id",
                "first_name",
                "last_name",
                "name",
                "title",
                "job_position",
                "email",
                "phone",
                "engagement_level",
                "contact_stage",
                "last_activity_date",
                "do_not_call",
            ],
            dataset.contacts,
            overwrite=False,
        ),
    ]
    return "\n\n".join(statement for statement in statements if statement)
